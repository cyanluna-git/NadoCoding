from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1) # 이동 범위 설정하고 이동 시킬 위치
# ws["B1"].value = "국어" # B1 셀에 국어 입력

#번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=1)

wb.save("sample_korean.xlsx")