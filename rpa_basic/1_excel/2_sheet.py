from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 기본이름으로 새로운 시트 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff0066" #RGB형태로

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("yourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"] # Dict 형태로 sheet 에 접근


# sheet복사
new_ws["A1"] = "Roy" # 데이터 넣기
new_ws["A2"] = "Gerald" # 데이터 넣기
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames) # 모든 시트이름 확인
wb.save("sample.xlsx")
