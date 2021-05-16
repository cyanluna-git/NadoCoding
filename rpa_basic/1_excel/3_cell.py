from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 'Apple'
ws["B2"] = 'Beta'
ws["B3"] = 'C'

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)

c = ws.cell(column=3, row=1, value='Meta')
print(c.value)

from random import  *
index = 1
for x in range(1, 11):
    for y in range(1,11):
        # ws.cell(row=x, column=y, value= randint(0, 100))
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
wb.close()